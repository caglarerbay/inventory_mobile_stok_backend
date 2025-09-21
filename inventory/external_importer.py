# inventory/external_importer.py

import openpyxl
import logging
import os
import tempfile
from django.db import transaction
from .models import ExternalProduct

from .models import Installation, DeviceRecord, Institution, DeviceType
from .models import Product

logger = logging.getLogger(__name__)

def run_import(file_path, batch_size=1000):
    logger.info(f"Import başlıyor: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    total = len(rows)

    # Excel'deki part_code'ları set olarak topla
    excel_codes = set(str(row[0]).strip() for row in rows if row[0])

    # DB'deki part_code'ları set olarak topla
    db_codes = set(ExternalProduct.objects.values_list('part_code', flat=True))

    # Excel’de olmayanları sil
    to_delete = db_codes - excel_codes
    if to_delete:
        ExternalProduct.objects.filter(part_code__in=to_delete).delete()

    for idx in range(0, total, batch_size):
        batch = rows[idx:idx+batch_size]
        batch_num = idx // batch_size + 1
        logger.info(f"[BATCH {batch_num}] İşleniyor: {len(batch)} kayıt")

        codes = [row[0] for row in batch]
        existing = ExternalProduct.objects.filter(part_code__in=codes)
        existing_map = {e.part_code: e for e in existing}

        to_create = []
        to_update = []

        for row in batch:
            code, name, devices, price = row
            if code in existing_map:
                obj = existing_map[code]
                obj.name = name
                obj.devices = devices
                obj.unit_price = price
                to_update.append(obj)
            else:
                to_create.append(
                    ExternalProduct(
                        part_code=code,
                        name=name,
                        devices=devices,
                        unit_price=price
                    )
                )

        with transaction.atomic():
            if to_create:
                ExternalProduct.objects.bulk_create(to_create, batch_size=batch_size)
            if to_update:
                ExternalProduct.objects.bulk_update(
                    to_update,
                    ['name', 'devices', 'unit_price'],
                    batch_size=batch_size
                )

        logger.info(f"[BATCH {batch_num}] Tamamlandı")

    logger.info("Import tamamlandı")
    try:
        os.remove(file_path)
    except OSError:
        pass

#kurumlar için parçalı import
def run_import_institutions(file_path, batch_size=1000):
    print(f"Import (Institution) başlıyor: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    total = len(rows)
    print(f"Toplam kurum satırı: {total}")

    # Excel'deki kurum isimlerini set olarak topla
    excel_names = set(str(row[0]).strip() for row in rows if row[0])

    # DB'deki kurum isimlerini set olarak topla
    db_names = set(Institution.objects.values_list('name', flat=True))

    # Excel’de olmayanları sil
    to_delete = db_names - excel_names
    if to_delete:
        Institution.objects.filter(name__in=to_delete).delete()

    for idx in range(0, total, batch_size):
        batch = rows[idx: idx + batch_size]
        batch_no = idx // batch_size + 1
        print(f"[INSTITUTION BATCH {batch_no}] İşleniyor: {len(batch)} kayıt")
        with transaction.atomic():
            for name, city, contact_name, contact_phone in batch:
                Institution.objects.update_or_create(
                    name=name,
                    defaults={
                        'city': city,
                        'contact_name': contact_name,
                        'contact_phone': contact_phone,
                    }
                )
        print(f"[INSTITUTION BATCH {batch_no}] Tamamlandı")

    print("Institution import tamamlandı")
    try:
        os.remove(file_path)
    except OSError:
        pass

#ürünler için batch import
def run_products_import(file_path, batch_size=1000):
    # 1) Excel dosyasını oku
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    # 2) Excel'deki part_code'ları topla
    excel_codes = [str(row[0]).strip() for row in rows if row[0] is not None]

    # 3) Veritabanındaki mevcut kayıtları çek
    existing = Product.objects.filter(part_code__in=excel_codes)
    existing_map = {e.part_code: e for e in existing}
    db_codes = list(existing_map.keys())

    to_create = []
    to_update = []

    # 4) Her satırı işle
    for row in rows:
        # Burada sütun sırası export sırasına göre:
        # part_code, name, cabinet, shelf, quantity, min_limit, order_placed
        part_code, name, cabinet, shelf, quantity, min_limit, order_placed = row

        if part_code is None:
            continue
        code = str(part_code).strip()

        defaults = {
            'name': name,
            'cabinet': cabinet,
            'shelf': shelf,
            'quantity': quantity or 0,
            'min_limit': min_limit or 0,
            'order_placed': bool(order_placed),
        }

        if code in existing_map:
            obj = existing_map[code]
            for k, v in defaults.items():
                setattr(obj, k, v)
            to_update.append(obj)
        else:
            to_create.append(Product(part_code=code, **defaults))

    # 5) Toplu kaydet ve silme
    with transaction.atomic():
        if to_create:
            Product.objects.bulk_create(to_create, batch_size=batch_size)
        if to_update:
            Product.objects.bulk_update(
                to_update,
                ['name', 'cabinet', 'shelf', 'quantity', 'min_limit', 'order_placed'],
                batch_size=batch_size
            )

        # Excel'de olmayanları veritabanından sil
        deleted_count, _ = Product.objects.exclude(part_code__in=excel_codes).delete()

    # 6) Geçici dosyayı sil
    try:
        os.remove(file_path)
    except OSError:
        pass

    # 7) Geri dön: excel kodları, DB kodları, silinen adedi
    return excel_codes, db_codes, deleted_count



#cihazlar için batch import
def run_device_records_import(file_path, batch_size=1000):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    # Excel'deki serial_number seti
    excel_serials = set(str(row[0]).strip() for row in rows if row[0])
    # DB'deki serial_number seti
    db_serials = set(DeviceRecord.objects.values_list('serial_number', flat=True))
    # Excel'de olmayanları sil
    to_delete = db_serials - excel_serials
    if to_delete:
        DeviceRecord.objects.filter(serial_number__in=to_delete).delete()

    for batch_index in range(0, len(rows), batch_size):
        batch_number = batch_index // batch_size + 1
        batch = rows[batch_index:batch_index + batch_size]
        _process_device_record_batch(batch, batch_number)

    try:
        os.remove(file_path)
    except OSError:
        pass

def _process_device_record_batch(batch, batch_number):
    print(f"[BATCH {batch_number}] İşleniyor: {len(batch)} kayıt")
    with transaction.atomic():
        for row in batch:
            serial_number, device_type_name, institution_name = row

            # DeviceType’ı name üzerinden buluyoruz
            try:
                device_type = DeviceType.objects.get(name=device_type_name)
            except DeviceType.DoesNotExist:
                raise Exception(f"DeviceType matching '{device_type_name}' bulunamadı.")

            # Institution varsa name üzerinden bul, yoksa None
            institution = None
            if institution_name and institution_name.strip():
                institution = Institution.objects.filter(name=institution_name).first()
                if not institution:
                    raise Exception(f"Institution matching '{institution_name}' bulunamadı.")

            DeviceRecord.objects.update_or_create(
                serial_number=serial_number,
                defaults={
                    'device_type': device_type,
                    'institution': institution,
                }
            )
    print(f"[BATCH {batch_number}] Tamamlandı")

#kurulumlar için batch import
def run_import_installations(file_path, batch_size=1000):
    import datetime
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    total = len(rows)
    print(f"Toplam {total} kayıt okunacak.")

    excel_keys = set()
    # İlk turda sadece anahtarları topla
    for row in rows:
        serial, inst_name, install_date, uninstall_date, core_serial = row
        try:
            device = DeviceRecord.objects.get(serial_number=serial)
        except DeviceRecord.DoesNotExist:
            continue
        # install_date bazen datetime.datetime olarak gelir, string yap
        key = (device.id, str(install_date)[:10])
        excel_keys.add(key)

    # DB'deki mevcut anahtarları çek (string ile karşılaştırmak için)
    db_keys = set(
        (inst.device_id, str(inst.install_date)[:10])
        for inst in Installation.objects.all()
    )

    # Excel'de olmayanları sil
    to_delete = db_keys - excel_keys
    if to_delete:
        for device_id, install_date in to_delete:
            Installation.objects.filter(device_id=device_id, install_date=install_date).delete()

    # Şimdi batch işlem ile ekle/güncelle
    def _process_batch(batch_rows, batch_index):
        print(f"[BATCH {batch_index+1}] İşleniyor: {len(batch_rows)} kayıt")
        for serial, inst_name, install_date, uninstall_date, core_serial in batch_rows:
            try:
                device = DeviceRecord.objects.get(serial_number=serial)
            except DeviceRecord.DoesNotExist:
                print(f"  !!! Cihaz bulunamadı: {serial}")
                continue

            try:
                institution = Institution.objects.get(name=inst_name)
            except Institution.DoesNotExist:
                print(f"  !!! Kurum bulunamadı: {inst_name}")
                continue

            connected_core = None
            if core_serial:
                connected_core = DeviceRecord.objects.filter(
                    serial_number=core_serial,
                    device_type__is_core=True
                ).first()
                if not connected_core:
                    print(f"  !!! Bağlı core bulunamadı veya core değil: {core_serial}")

            Installation.objects.update_or_create(
                device=device,
                install_date=install_date,
                defaults={
                    'institution': institution,
                    'uninstall_date': uninstall_date or None,
                    'connected_core': connected_core,
                }
            )
        print(f"[BATCH {batch_index+1}] Tamamlandı")

    with transaction.atomic():
        for i in range(0, total, batch_size):
            batch = rows[i:i+batch_size]
            _process_batch(batch, i // batch_size)

    print("Import tamamlandı.")
    try:
        os.remove(file_path)
    except OSError:
        pass


    # toplu işlem
    with transaction.atomic():
        for i in range(0, total, batch_size):
            batch = rows[i:i+batch_size]
            _process_batch(batch, i // batch_size)

    print("Import tamamlandı.")
    # istersen geçici dosyayı sil
    try:
        os.remove(file_path)
    except OSError:
        pass